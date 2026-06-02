from __future__ import annotations
"""CSV + Excel delivery.

Each batch produces two files:
  1. leads_batch_N_YYYYMMDD_HHMM.csv  — flat CSV, all leads, one row per lead
  2. leads_batch_N_YYYYMMDD_HHMM.xlsx — Excel workbook with:
       • "All Leads" sheet  — same as the CSV
       • One sheet per niche (e.g. "Marketing Agency", "Coaching", "SaaS / Tech")
         containing only leads that belong to that niche

The hook column (personalization icebreaker from Claude) is included in both formats.
"""
import csv
import os
import smtplib
import logging
from collections import defaultdict
from email.message import EmailMessage
from datetime import datetime
from sqlalchemy import select, func, or_
from db import SessionLocal, VerifiedLead, Batch
from config import settings

log = logging.getLogger("delivery")


def _deliverable_filter():
    """SQLAlchemy filter for leads that are SAFE to export/deliver.

    Excludes:
      • bounced leads (confirmed dead by Instantly) — the #1 source of the
        "download all gives bad results" complaint.
      • catch-all leads when EXPORT_CATCH_ALL is off — catch-all domains accept
        any address at SMTP time, so they're a major silent-bounce risk.

    NULL-safe on purpose: older rows may have NULL in bounced / is_catch_all.
    We treat NULL as "not bounced" / "not catch-all" so we never drop a clean
    legacy lead, while still hard-excluding rows explicitly flagged True.
    """
    conds = [or_(VerifiedLead.bounced == False, VerifiedLead.bounced.is_(None))]
    if not settings.EXPORT_CATCH_ALL:
        conds.append(or_(VerifiedLead.is_catch_all == False,
                         VerifiedLead.is_catch_all.is_(None)))
    return conds

# Column order for export.
# Headers are whitespace-free so Instantly (and other tools) import without errors.
# Instantly auto-maps "firstName", "lastName", "email" — keep those exact names.
COLUMNS = [
    "email", "firstName", "lastName", "role", "niche",
    "company", "website", "tier", "source", "sourceUrl",
    "verifyStatus", "score", "catchAll", "hook",
]


def _clean(val) -> str:
    """Strip ALL whitespace variants (including Unicode) that break Instantly."""
    import re as _re, unicodedata as _ud
    v = str(val or "")
    # Remove zero-width / invisible Unicode chars (U+200B, U+FEFF, U+00AD etc.)
    v = _re.sub(r'[­​‌‍‎‏‪-‮⁠﻿]', '', v)
    # Collapse newlines/tabs to space
    v = _re.sub(r'[\r\n\t]+', ' ', v)
    # Strip C0/C1 control chars
    v = ''.join(c for c in v if _ud.category(c) != 'Cc')
    # Collapse multiple spaces
    v = _re.sub(r' {2,}', ' ', v)
    return v.strip()


def _lead_row(l: VerifiedLead) -> list:
    return [
        _clean(l.email),
        _clean(l.first_name),
        _clean(l.last_name),
        _clean(l.role),
        _clean(l.niche or "Founder / Startup"),
        _clean(l.company),
        _clean(l.website),
        l.tier or "A",
        _clean(l.source),
        _clean(l.source_url),
        _clean(l.reoon_status),
        l.reoon_score or "",
        "Yes" if l.is_catch_all else "No",
        _clean(l.hook),
    ]


def _write_csv(path: str, leads: list[VerifiedLead]) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(COLUMNS)
        for l in leads:
            w.writerow(_lead_row(l))


def _write_excel(path: str, leads: list[VerifiedLead]) -> None:
    """Write multi-sheet Excel. Sheet 'All Leads' + one per niche."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.warning("openpyxl not installed — skipping Excel output")
        return

    wb = Workbook()

    HEADER_FILL = PatternFill("solid", fgColor="1a1a2e")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
    ALT_FILL = PatternFill("solid", fgColor="F0F4FF")

    def _add_sheet(ws, sheet_leads: list[VerifiedLead]):
        ws.append(COLUMNS)
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        for i, lead in enumerate(sheet_leads, start=2):
            ws.append(_lead_row(lead))
            if i % 2 == 0:
                for cell in ws[i]:
                    cell.fill = ALT_FILL

        # Auto-fit columns (approximate)
        col_widths = {c: len(c) + 2 for c in COLUMNS}
        for lead in sheet_leads[:200]:  # sample first 200 rows for width calc
            for j, val in enumerate(_lead_row(lead)):
                col_widths[COLUMNS[j]] = min(60, max(col_widths[COLUMNS[j]], len(str(val)) + 2))
        for j, col in enumerate(COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(j)].width = col_widths[col]

    # All Leads sheet
    ws_all = wb.active
    ws_all.title = "All Leads"
    _add_sheet(ws_all, leads)

    # Per-niche sheets — skip niches with < 2 leads (not worth a separate tab)
    by_niche: dict[str, list[VerifiedLead]] = defaultdict(list)
    for l in leads:
        niche = (l.niche or "Founder / Startup").strip() or "Founder / Startup"
        by_niche[niche].append(l)

    for niche in sorted(by_niche.keys(), key=lambda n: -len(by_niche[n])):
        niche_leads = by_niche[niche]
        if len(niche_leads) < 2:
            continue
        # Excel sheet names: max 31 chars, no invalid chars
        sheet_name = niche[:31].replace("/", "-").replace("\\", "-").replace("*", "").replace("?", "").replace("[", "").replace("]", "").replace(":", "")
        ws = wb.create_sheet(title=sheet_name)
        _add_sheet(ws, niche_leads)

    wb.save(path)
    log.info(f"Excel written: {path} ({len(leads)} leads, {len(by_niche)} niches)")


async def deliver_all_leads_csv() -> str:
    """Write a CSV of ALL verified leads ever — useful for one-shot downloads."""
    async with SessionLocal() as s:
        result = await s.execute(
            select(VerifiedLead).where(*_deliverable_filter()).order_by(VerifiedLead.id)
        )
        leads = result.scalars().all()

    fname = f"all_leads_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.csv"
    path = os.path.join(settings.DATA_DIR, fname)
    _write_csv(path, leads)
    return path


async def deliver_leads_range(start: int, end: int) -> tuple[str, int, int]:
    """Write a CSV of clean leads numbered [start..end] (1-indexed, inclusive).

    Numbering is over the DELIVERABLE set only (bounced + catch-all already
    excluded), ordered oldest-first by id so row numbers stay stable across
    downloads. Example: start=8001, end=10000 → the next 2000 good leads after
    the first 8000 you already pulled.

    Returns (csv_path, rows_written, total_available).
    """
    start = max(1, int(start))
    end = max(start, int(end))
    offset = start - 1
    limit = end - start + 1

    async with SessionLocal() as s:
        total_available = (await s.execute(
            select(func.count()).select_from(VerifiedLead).where(*_deliverable_filter())
        )).scalar_one()
        result = await s.execute(
            select(VerifiedLead)
            .where(*_deliverable_filter())
            .order_by(VerifiedLead.id)
            .offset(offset)
            .limit(limit)
        )
        leads = result.scalars().all()

    fname = f"leads_{start}-{end}_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.csv"
    path = os.path.join(settings.DATA_DIR, fname)
    _write_csv(path, leads)
    return path, len(leads), total_available


async def regenerate_csv_for_batch(batch_id: int) -> str | None:
    async with SessionLocal() as s:
        result = await s.execute(
            select(VerifiedLead).where(VerifiedLead.batch_id == batch_id, *_deliverable_filter())
        )
        leads = result.scalars().all()
        if not leads:
            return None
        b = (await s.execute(select(Batch).where(Batch.id == batch_id))).scalar_one_or_none()

    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M")
    fname = f"leads_batch_{batch_id}_{timestamp}.csv"
    path = os.path.join(settings.DATA_DIR, fname)
    _write_csv(path, leads)

    if b:
        from sqlalchemy import update as sql_update
        async with SessionLocal() as s:
            await s.execute(sql_update(Batch).where(Batch.id == batch_id).values(
                csv_path=path, delivered_count=len(leads),
            ))
            await s.commit()
    return path


async def deliver_batch(batch_id: int) -> str:
    """Write CSV + Excel for batch, attempt email delivery. Return CSV path."""
    async with SessionLocal() as s:
        result = await s.execute(
            select(VerifiedLead).where(VerifiedLead.batch_id == batch_id, *_deliverable_filter())
        )
        leads = result.scalars().all()

    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M")
    base_name = f"leads_batch_{batch_id}_{timestamp}"
    csv_path = os.path.join(settings.DATA_DIR, base_name + ".csv")
    xlsx_path = os.path.join(settings.DATA_DIR, base_name + ".xlsx")

    _write_csv(csv_path, leads)
    _write_excel(xlsx_path, leads)
    log.info(f"Batch {batch_id}: {len(leads)} leads → CSV + Excel")

    # Niche breakdown for the email body
    by_niche: dict[str, int] = defaultdict(int)
    for l in leads:
        by_niche[(l.niche or "Founder / Startup")] += 1

    try:
        if settings.SMTP_HOST and settings.SMTP_USER and settings.SMTP_PASS:
            attachments = []
            if leads:
                attachments.append(csv_path)
                if os.path.exists(xlsx_path):
                    attachments.append(xlsx_path)
            send_email(
                to=settings.DELIVERY_EMAIL,
                subject=f"[Lead Engine] Batch #{batch_id} — {len(leads)} verified leads ({len(by_niche)} niches)",
                body=build_email_body(batch_id, leads, csv_path, by_niche),
                attachments=attachments,
            )
            log.info(f"Email sent to {settings.DELIVERY_EMAIL}")
    except Exception as e:
        log.exception(f"Email delivery failed: {e}")

    return csv_path


def build_email_body(batch_id: int, leads, path: str, by_niche: dict) -> str:
    tier_a = sum(1 for l in leads if l.tier == "A")
    niche_lines = "\n".join(
        f"  • {n}: {c}" for n, c in sorted(by_niche.items(), key=lambda x: -x[1])
    )
    src_counts: dict[str, int] = {}
    for l in leads:
        src_counts[l.source or "Other"] = src_counts.get(l.source or "Other", 0) + 1
    src_lines = "\n".join(f"  • {s}: {c}" for s, c in sorted(src_counts.items(), key=lambda x: -x[1]))

    download_link = ""
    if settings.PUBLIC_BASE_URL:
        download_link = f"\nDownload CSV: {settings.PUBLIC_BASE_URL}/download/{os.path.basename(path)}\n"

    hook_count = sum(1 for l in leads if l.hook)

    return f"""Hey,

Batch #{batch_id} is ready.

Tier-A verified leads: {tier_a}
Leads with personalisation hooks (from Claude): {hook_count}
{download_link}
Niche breakdown:
{niche_lines}

Source breakdown:
{src_lines}

The Excel attachment has one sheet per niche — use the Hook column for personalised outreach openers.

— Lead Engine
"""


async def notify_sources_exhausted():
    """Email when source pool is dry — ask for direction."""
    async with SessionLocal() as s:
        total = (await s.execute(select(func.count()).select_from(VerifiedLead))).scalar_one()
        last_batches = (await s.execute(
            select(Batch).order_by(Batch.id.desc()).limit(5)
        )).scalars().all()

    body = f"""Hey,

The lead engine has worked through every URL in the current source pool.

Running total Tier-A leads delivered: {total}

Recent batches:
""" + "\n".join(f"  • Batch #{b.id}: {b.delivered_count or 0} leads ({b.trigger})" for b in last_batches) + """

Next steps:
  (a) New sources: reply with a site URL to add
  (b) Seed URLs: paste a list of profile/interview URLs
  (c) Pause: POST /pause
  (d) Wait: sources publish new content weekly

— Lead Engine
"""
    try:
        if settings.SMTP_HOST and settings.SMTP_USER and settings.SMTP_PASS:
            send_email(
                to=settings.DELIVERY_EMAIL,
                subject="[Lead Engine] Sources exhausted — need direction",
                body=body, attachments=[],
            )
    except Exception as e:
        log.exception(f"notify_sources_exhausted email failed: {e}")


def send_email(to: str, subject: str, body: str, attachments: list[str] = None):
    msg = EmailMessage()
    msg["From"] = settings.SMTP_USER
    msg["To"] = to
    msg["Subject"] = subject
    msg.set_content(body)

    for ap in (attachments or []):
        if not os.path.exists(ap):
            continue
        with open(ap, "rb") as f:
            data = f.read()
        # Pick correct MIME subtype
        if ap.endswith(".xlsx"):
            msg.add_attachment(
                data,
                maintype="application",
                subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=os.path.basename(ap),
            )
        else:
            msg.add_attachment(
                data, maintype="text", subtype="csv",
                filename=os.path.basename(ap),
            )

    with smtplib.SMTP(settings.SMTP_HOST, settings.SMTP_PORT) as smtp:
        smtp.starttls()
        smtp.login(settings.SMTP_USER, settings.SMTP_PASS)
        smtp.send_message(msg)
